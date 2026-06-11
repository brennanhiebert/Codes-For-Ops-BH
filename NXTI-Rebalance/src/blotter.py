"""
blotter.py — assemble trade legs into the final blotter, compute realized P&L,
and write a formatted multi-sheet workbook (Blotter / Summary / Exceptions).

xlsx formatting conventions (per the xlsx skill):
  * negatives shown in parentheses
  * currency as $#,##0
  * zero rendered as a dash
We write static VALUES (no formulas), so there are zero formula errors by design.
"""
from __future__ import annotations

import pandas as pd

from . import config, lot_engine
from .lot_engine import Lot

# Number formats (value;negative-in-parens;zero-as-dash).
FMT_USD = '$#,##0;($#,##0);"-"'
FMT_USD2 = '$#,##0.00;($#,##0.00);"-"'
FMT_SHARES = '#,##0;(#,##0);"-"'
FMT_PRICE = '$#,##0.0000;($#,##0.0000);"-"'

BLOTTER_COLUMNS = [
    "cusip", "isin", "security_name", "classification",
    "current_shares", "target_shares", "side", "method",
    "shares", "current_price", "notional", "lots_used", "multi_lot",
    "avg_basis", "realized_gain_loss", "lot_detail", "flags",
]


def _save(wb, path):
    """
    Save, but if the target file is open in Excel (PermissionError) fall back to a
    timestamped filename so a run never fails just because a file is open.
    Returns the path actually written.
    """
    import datetime as _dt

    try:
        wb.save(path)
        return path
    except PermissionError:
        alt = path.with_name(
            f"{path.stem}_{_dt.datetime.now():%H%M%S}{path.suffix}"
        )
        wb.save(alt)
        import warnings
        warnings.warn(
            f"'{path.name}' is open/locked (close it in Excel). "
            f"Wrote to '{alt.name}' instead."
        )
        return alt


def _save_csv(df: pd.DataFrame, path):
    """to_csv with the same open-in-Excel fallback as _save()."""
    import datetime as _dt

    try:
        df.to_csv(path, index=False)
        return path
    except PermissionError:
        alt = path.with_name(
            f"{path.stem}_{_dt.datetime.now():%H%M%S}{path.suffix}"
        )
        df.to_csv(alt, index=False)
        import warnings
        warnings.warn(
            f"'{path.name}' is open/locked (close it in Excel). "
            f"Wrote to '{alt.name}' instead."
        )
        return alt


def _lot_detail_str(fills) -> str:
    return "; ".join(
        f"{f.open_date}:{f.shares:g}@{f.basis:.4f}{'*' if f.was_split else ''}"
        for f in fills
    )


def assemble_blotter(
    delta: pd.DataFrame,
    lots_df: pd.DataFrame,
    recon: pd.DataFrame,
) -> tuple[pd.DataFrame, dict, pd.DataFrame]:
    """Returns (blotter_df, summary_dict, exceptions_df)."""
    lots_by_cusip: dict[str, list[Lot]] = {}
    for cusip, g in lots_df.groupby("cusip"):
        lots_by_cusip[cusip] = [
            Lot(open_date=str(r.open_date), shares=float(r.shares), basis=float(r.basis))
            for r in g.itertuples()
        ]
    recon_map = recon.set_index("cusip").to_dict("index")

    rows: list[dict] = []
    records: list[dict] = []   # same as rows, but each carries its raw lot fills
    exceptions: list[dict] = []

    for r in delta.itertuples():
        base = {
            "cusip": r.cusip,
            "isin": r.isin,
            "security_name": r.security_name,
            "classification": r.classification,
            "current_shares": r.current_qty,
            "target_shares": r.target_shares,
            "current_price": r.current_price,
        }

        if r.action == "NO_TRADE":
            # Not a blotter leg, but carried in `records` so the EXPANDABLE
            # workbook can show the full universe (at-target names included).
            records.append({
                **base, "side": "NO_TRADE", "method": "NO_TRADE",
                "shares": 0.0, "notional": 0.0,
                "lots_used": 0, "multi_lot": False,
                "avg_basis": float("nan"), "realized_gain_loss": 0.0,
                "lot_detail": "", "flags": "at target (within rounding)",
                "fills": [],
            })
            continue

        if r.action == "BUY":
            shares = r.qty_to_cover
            notional = shares * r.current_price
            flags = []
            if pd.isna(r.current_price):
                flags.append("missing price")
            row = {
                **base, "side": "BUY", "method": "MARKET",
                "shares": shares, "notional": notional,
                "lots_used": 0, "multi_lot": False,
                "avg_basis": float("nan"), "realized_gain_loss": 0.0,
                "lot_detail": "", "flags": "; ".join(flags),
            }
            rows.append(row)
            records.append({**row, "fills": []})
            if flags:
                exceptions.append({**base, "side": "BUY", "issue": "; ".join(flags)})
            continue

        # ---- SELL (FULL or PARTIAL) ----
        is_full = (r.action == "FULL_SELL")
        lots = lots_by_cusip.get(r.cusip, [])
        legs = lot_engine.build_sell_legs(
            lots=lots,
            current_price=r.current_price,
            qty_to_cover=r.qty_to_cover,
            is_full_sell=is_full,
        )

        # Reconciliation flag: do lots tie to the summary row / tracker qty?
        rc = recon_map.get(r.cusip)
        recon_flag = ""
        if rc is not None and not rc["reconciles"]:
            recon_flag = "share reconciliation mismatch"
        if is_full and rc is not None:
            # For a full sell we dispose every lot; warn if lot total != held qty.
            if abs(rc["lot_shares"] - r.current_qty) > config.TRADE_EPSILON:
                recon_flag = (recon_flag + "; " if recon_flag else "") + (
                    f"lot total {rc['lot_shares']:g} != held qty {r.current_qty:g}"
                )

        for leg in legs:
            flags = list(leg.flags)
            if recon_flag:
                flags.append(recon_flag)
            rgl = leg.realized_gain_loss(r.current_price)
            row = {
                **base, "side": "SELL", "method": leg.method,
                "shares": leg.shares, "notional": leg.shares * r.current_price,
                "lots_used": leg.lots_used, "multi_lot": leg.multi_lot,
                "avg_basis": leg.avg_basis,
                "realized_gain_loss": rgl if leg.method == "MARKET" else 0.0,
                "lot_detail": _lot_detail_str(leg.fills),
                "flags": "; ".join(flags),
            }
            rows.append(row)
            records.append({**row, "fills": list(leg.fills)})
            if leg.method == "UNDETERMINED" or any(
                k in (row["flags"] or "")
                for k in ("no tax lots", "shortfall", "reconciliation", "!= held")
            ):
                exceptions.append({
                    **base, "side": "SELL", "method": leg.method,
                    "shares": leg.shares, "issue": row["flags"],
                })

    blotter = pd.DataFrame(rows, columns=BLOTTER_COLUMNS)
    exceptions_df = pd.DataFrame(exceptions)
    summary = _summarize(delta, blotter)
    return blotter, summary, exceptions_df, records


def _summarize(delta: pd.DataFrame, blotter: pd.DataFrame) -> dict:
    cls = delta["classification"].value_counts().to_dict()
    cont = delta[delta["classification"] == "CONTINUING"]["action"].value_counts().to_dict()

    buys = blotter[blotter["side"] == "BUY"]
    sells = blotter[blotter["side"] == "SELL"]
    mkt = sells[sells["method"] == "MARKET"]
    ink = sells[sells["method"] == "IN_KIND"]

    buy_notional = float(buys["notional"].sum())
    mkt_notional = float(mkt["notional"].sum())
    ink_notional = float(ink["notional"].sum())
    realized = blotter["realized_gain_loss"].fillna(0.0)
    realized_loss = float(realized[realized < 0].sum())
    realized_gain = float(realized[realized > 0].sum())

    # Embedded gain SHIELDED by in-kind transfer (never realized — it leaves
    # with the shares): (snapshot price - avg basis) * shares per in-kind leg.
    shielded_gain = float(
        ((ink["current_price"] - ink["avg_basis"]) * ink["shares"])
        .fillna(0.0).sum()
    )

    # Net cash impact: buys consume cash, market sells raise cash; in-kind is cashless.
    net_cash = mkt_notional - buy_notional

    return {
        "Run date": config.RUN_DATE,
        "Price source": "",  # filled by caller
        "Partial-sell priority": config.PARTIAL_SELL_PRIORITY,
        "Investable base": config.INVESTABLE_BASE,
        "—": "",
        "ADDs": cls.get("ADD", 0),
        "DROPs": cls.get("DROP", 0),
        "CONTINUING": cls.get("CONTINUING", 0),
        "  CONTINUING buys": cont.get("BUY", 0),
        "  CONTINUING partial sells": cont.get("PARTIAL_SELL", 0),
        "  CONTINUING no-trade": cont.get("NO_TRADE", 0),
        "——": "",
        "Buy legs": int(len(buys)),
        "Market sell legs": int(len(mkt)),
        "In-kind sell legs": int(len(ink)),
        "———": "",
        "Total buy notional": buy_notional,
        "Total market sell notional": mkt_notional,
        "Total in-kind notional": ink_notional,
        "————": "",
        "Total realized loss (harvested)": realized_loss,
        "Total realized gain (should be ~0)": realized_gain,
        "Net realized P&L": realized_loss + realized_gain,
        "Embedded gain shielded (in-kind)": shielded_gain,
        "—————": "",
        "Net cash impact (mkt sells - buys)": net_cash,
    }


# --------------------------------------------------------------------------- #
# AP in-kind instruction file
# --------------------------------------------------------------------------- #
def write_ap_inkind_csv(blotter_df: pd.DataFrame, tracker: pd.DataFrame, path):
    """
    The in-kind instruction file sent to BNY/the AP: one row per security to be
    transferred in-kind, identifiers + share quantity.

    SHARES ARE THE AUTHORITATIVE FIELD. indicative_price / indicative_notional
    are valued off this run's price snapshot and are included for sizing context
    only — the AP transacts the share quantities, not the notionals.

    Tickers come from the tracker (plain exchange ticker, e.g. "AFL") — every
    in-kind name is a held position, so coverage is complete.
    """
    ink = blotter_df[
        (blotter_df["side"] == "SELL") & (blotter_df["method"] == "IN_KIND")
    ].copy()

    ticker = tracker.set_index("cusip")["ticker"]
    ink["ticker"] = ink["cusip"].map(ticker)
    ink["shares"] = ink["shares"].round(4)  # strip float noise; lots can be fractional
    ink["indicative_price"] = ink["current_price"]
    ink["indicative_notional"] = (ink["shares"] * ink["current_price"]).round(2)

    out = ink[["ticker", "cusip", "isin", "security_name", "shares",
               "indicative_price", "indicative_notional"]] \
        .sort_values("indicative_notional", ascending=False) \
        .reset_index(drop=True)
    path = _save_csv(out, path)

    print(f"[ap] in-kind instruction file -> {path.name}: "
          f"{len(out)} securities, {out['shares'].sum():,.0f} shares, "
          f"~${out['indicative_notional'].sum():,.0f} indicative notional")
    return path


# --------------------------------------------------------------------------- #
# Trader market-execution file
# --------------------------------------------------------------------------- #
def write_trader_csv(blotter_df: pd.DataFrame, tracker: pd.DataFrame,
                     proforma: pd.DataFrame, path):
    """
    The market-execution list for the trading desk: every BUY and every MARKET
    SELL, one row per security per side. IN_KIND legs are excluded — those go
    to the AP via the in-kind instruction file, not to the desk.

    SHARES ARE THE AUTHORITATIVE FIELD. indicative_price / indicative_notional
    are valued off this run's price snapshot for sizing context only.
    """
    mkt = blotter_df[
        (blotter_df["side"] == "BUY")
        | ((blotter_df["side"] == "SELL") & (blotter_df["method"] == "MARKET"))
    ].copy()

    # Plain ticker: tracker for held names; ADDs aren't held, so fall back to
    # the pro-forma's Bloomberg ticker ("FTNT UW Equity" -> "FTNT").
    ticker = tracker.set_index("cusip")["ticker"]
    pf_ticker = (proforma.set_index("cusip")["ticker"]
                 .str.split().str[0])
    mkt["ticker"] = mkt["cusip"].map(ticker).fillna(mkt["cusip"].map(pf_ticker))
    mkt["shares"] = mkt["shares"].round(4)
    mkt["indicative_price"] = mkt["current_price"]
    mkt["indicative_notional"] = (mkt["shares"] * mkt["current_price"]).round(2)

    out = mkt[["side", "ticker", "cusip", "isin", "security_name", "shares",
               "indicative_price", "indicative_notional"]] \
        .sort_values(["side", "indicative_notional"], ascending=[True, False]) \
        .reset_index(drop=True)
    path = _save_csv(out, path)

    n_buy = int((out["side"] == "BUY").sum())
    n_sell = int((out["side"] == "SELL").sum())
    buy_notional = out.loc[out["side"] == "BUY", "indicative_notional"].sum()
    sell_notional = out.loc[out["side"] == "SELL", "indicative_notional"].sum()
    print(f"[desk] trader execution file -> {path.name}: "
          f"{n_buy} buys (~${buy_notional:,.0f}), "
          f"{n_sell} market sells (~${sell_notional:,.0f})")
    return path


# --------------------------------------------------------------------------- #
# Custodian sell-method file
# --------------------------------------------------------------------------- #
def write_custodian_csv(blotter_df: pd.DataFrame, tracker: pd.DataFrame, path):
    """
    Custodian accounting instruction file: ONE row per security being sold,
    stating whether the ENTIRE reduction is an IN_KIND transfer or a MARKET
    sell, so the custodian applies the right back-end accounting. Lot relief
    is intentionally NOT specified — BNY selects the lots on their end.

    One row per security is guaranteed by SELL_METHOD_POLICY="all_or_nothing";
    under the legacy split policy a security can emit one row per method.
    """
    sells = blotter_df[blotter_df["side"] == "SELL"].copy()

    ticker = tracker.set_index("cusip")["ticker"]
    sells["ticker"] = sells["cusip"].map(ticker)
    sells["shares"] = sells["shares"].round(4)

    out = sells[["ticker", "cusip", "isin", "security_name", "side",
                 "method", "shares"]] \
        .sort_values(["method", "security_name"]).reset_index(drop=True)
    path = _save_csv(out, path)

    by_method = out["method"].value_counts().to_dict()
    n_dupe = int(out["cusip"].duplicated().sum())
    print(f"[cust] sell-method file -> {path.name}: {len(out)} securities "
          f"(by method: {by_method})"
          f"{'; %d MULTI-METHOD security!' % n_dupe if n_dupe else ''}")
    return path


# --------------------------------------------------------------------------- #
# Excel writer
# --------------------------------------------------------------------------- #
def write_xlsx(
    blotter: pd.DataFrame,
    summary: dict,
    exceptions: pd.DataFrame,
    path,
) -> None:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Palette
    NAVY = "1F3864"; BLUE = "2E5496"; LIGHT = "D9E1F2"; GREY = "F2F2F2"
    GREEN = "C6EFCE"; RED = "FFC7CE"; AMBER = "FFEB9C"
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor=NAVY)
    title_font = Font(bold=True, color="FFFFFF", size=14)
    title_fill = PatternFill("solid", fgColor=BLUE)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # ----------------------------------------------------------------- #
    # SUMMARY sheet
    # ----------------------------------------------------------------- #
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:C1")
    ws["A1"] = f"NXTI REBALANCE - TRADE BLOTTER SUMMARY  ({config.RUN_DATE})"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 26

    money_keys = {
        "Total buy notional", "Total market sell notional", "Total in-kind notional",
        "Total realized loss (harvested)", "Total realized gain (should be ~0)",
        "Net realized P&L", "Net cash impact (mkt sells - buys)",
        "Embedded gain shielded (in-kind)",
    }
    r = 3
    for k, v in summary.items():
        if str(k).strip().startswith("—") or k == "—":  # separator rows
            r += 1
            continue
        c_label = ws.cell(row=r, column=1, value=k)
        c_val = ws.cell(row=r, column=2, value=v)
        c_label.font = Font(bold=k.endswith(":") or not k.startswith("  "))
        c_label.alignment = left
        if k in money_keys:
            c_val.number_format = FMT_USD
            if "loss" in k.lower() and isinstance(v, (int, float)) and v < 0:
                c_val.fill = PatternFill("solid", fgColor=GREEN)  # harvested loss = good
            if "should be ~0" in k and isinstance(v, (int, float)) and abs(v) > 1:
                c_val.fill = PatternFill("solid", fgColor=RED)
        elif isinstance(v, (int, float)):
            c_val.number_format = FMT_SHARES
        c_val.alignment = Alignment(horizontal="right")
        r += 1
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 20

    # Mini bar chart: notional by leg type.
    chart_anchor_row = r + 2
    ws.cell(row=chart_anchor_row, column=1, value="Notional by type").font = Font(bold=True)
    data_start = chart_anchor_row + 1
    chart_rows = [
        ("Buys", abs(summary["Total buy notional"])),
        ("Market sells", abs(summary["Total market sell notional"])),
        ("In-kind", abs(summary["Total in-kind notional"])),
    ]
    for i, (lab, val) in enumerate(chart_rows):
        ws.cell(row=data_start + i, column=1, value=lab)
        c = ws.cell(row=data_start + i, column=2, value=val)
        c.number_format = FMT_USD
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Trade Notional by Type"
    chart.legend = None
    chart.height = 5.5
    chart.width = 12
    data = Reference(ws, min_col=2, min_row=data_start, max_row=data_start + 2)
    cats = Reference(ws, min_col=1, min_row=data_start, max_row=data_start + 2)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    ws.add_chart(chart, f"D3")

    # ----------------------------------------------------------------- #
    # BLOTTER sheet
    # ----------------------------------------------------------------- #
    wsb = wb.create_sheet("Blotter")
    wsb.sheet_view.showGridLines = False
    headers = [c.replace("_", " ").title() for c in BLOTTER_COLUMNS]
    for j, h in enumerate(headers, start=1):
        c = wsb.cell(row=1, column=j, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = border

    money_cols = {"notional", "realized_gain_loss"}
    price_cols = {"current_price", "avg_basis"}
    share_cols = {"shares", "current_shares", "target_shares"}
    for i, row in enumerate(blotter.itertuples(index=False), start=2):
        rowd = dict(zip(BLOTTER_COLUMNS, row))
        for j, col in enumerate(BLOTTER_COLUMNS, start=1):
            val = rowd[col]
            if isinstance(val, float) and pd.isna(val):
                val = None
            c = wsb.cell(row=i, column=j, value=val)
            c.border = border
            if col in money_cols:
                c.number_format = FMT_USD
            elif col in price_cols:
                c.number_format = FMT_PRICE
            elif col in share_cols:
                c.number_format = FMT_SHARES
        # Row coloring by side/method.
        side = rowd["side"]; method = rowd["method"]
        if side == "BUY":
            fill = PatternFill("solid", fgColor=LIGHT)
        elif method == "MARKET":
            fill = PatternFill("solid", fgColor=GREEN)   # market sell = harvest loss
        elif method == "IN_KIND":
            fill = PatternFill("solid", fgColor=GREY)
        else:
            fill = PatternFill("solid", fgColor=AMBER)   # UNDETERMINED
        for j in range(1, len(BLOTTER_COLUMNS) + 1):
            wsb.cell(row=i, column=j).fill = fill
        if rowd["flags"]:
            wsb.cell(row=i, column=BLOTTER_COLUMNS.index("flags") + 1).fill = \
                PatternFill("solid", fgColor=AMBER)

    wsb.freeze_panes = "A2"
    wsb.auto_filter.ref = f"A1:{get_column_letter(len(BLOTTER_COLUMNS))}{len(blotter) + 1}"
    widths = {
        "cusip": 12, "isin": 14, "security_name": 32, "classification": 13,
        "current_shares": 13, "target_shares": 13,
        "side": 6, "method": 12, "shares": 11, "current_price": 13,
        "notional": 15, "lots_used": 9, "multi_lot": 9, "avg_basis": 12,
        "realized_gain_loss": 16, "lot_detail": 55, "flags": 30,
    }
    for j, col in enumerate(BLOTTER_COLUMNS, start=1):
        wsb.column_dimensions[get_column_letter(j)].width = widths.get(col, 12)

    # ----------------------------------------------------------------- #
    # EXCEPTIONS sheet
    # ----------------------------------------------------------------- #
    wse = wb.create_sheet("Exceptions")
    wse.sheet_view.showGridLines = False
    if exceptions is None or exceptions.empty:
        wse["A1"] = "No exceptions flagged. ✅"
        wse["A1"].font = Font(bold=True, size=12, color="006100")
        wse["A1"].fill = PatternFill("solid", fgColor=GREEN)
    else:
        ecols = list(exceptions.columns)
        for j, h in enumerate(ecols, start=1):
            c = wse.cell(row=1, column=j, value=h.replace("_", " ").title())
            c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = border
        for i, row in enumerate(exceptions.itertuples(index=False), start=2):
            for j, val in enumerate(row, start=1):
                if isinstance(val, float) and pd.isna(val):
                    val = None
                c = wse.cell(row=i, column=j, value=val)
                c.border = border
                c.fill = PatternFill("solid", fgColor=RED)
        wse.freeze_panes = "A2"
        for j, col in enumerate(ecols, start=1):
            wse.column_dimensions[get_column_letter(j)].width = \
                40 if col == "issue" else 18

    # Sheet order: Summary, Blotter, Exceptions
    wb.move_sheet("Summary", -wb.index(wb["Summary"]))
    return _save(wb, path)


def write_expandable_xlsx(records: list[dict], path) -> None:
    """
    A separate workbook where every trade leg is a PARENT row and the individual
    tax lots consumed sit GROUPED + COLLAPSIBLE beneath it (Excel's native +/-
    outline buttons in the left margin). Click '+' on a sell leg to reveal each
    lot: open date, shares, that lot's basis, today's price, the per-share P&L,
    and the lot's realized loss (MARKET) or gain shielded (IN_KIND).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    NAVY = "1F3864"; BLUE = "2E5496"; LIGHT = "D9E1F2"
    GREEN = "C6EFCE"; GREY = "EDEDED"; AMBER = "FFEB9C"; SUBHDR = "8EA9DB"
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor=NAVY)
    title_font = Font(bold=True, color="FFFFFF", size=14)
    title_fill = PatternFill("solid", fgColor=BLUE)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    italic = Font(italic=True, color="595959")

    wb = Workbook()

    # ----------------------------------------------------------------- #
    # SUMMARY sheet — tax-loss harvest vs gain shielded, lot-precise
    # ----------------------------------------------------------------- #
    mkt_legs = [r for r in records if r["side"] == "SELL" and r["method"] == "MARKET"]
    ink_legs = [r for r in records if r["side"] == "SELL" and r["method"] == "IN_KIND"]
    harvested = sum(r["realized_gain_loss"] for r in mkt_legs)
    shielded = sum(
        (r["current_price"] - f.basis) * f.shares
        for r in ink_legs for f in (r.get("fills") or [])
    )

    wss = wb.active
    wss.title = "Summary"
    wss.sheet_view.showGridLines = False
    wss.merge_cells("A1:C1")
    wss["A1"] = f"NXTI REBALANCE - TAX OUTCOME SUMMARY  ({config.RUN_DATE})"
    wss["A1"].font = title_font
    wss["A1"].fill = title_fill
    wss["A1"].alignment = center
    wss.row_dimensions[1].height = 26

    summary_rows = [
        ("TAX LOSSES HARVESTED (sold at market)", harvested, GREEN),
        (f"    across {len(mkt_legs)} market-sell legs / "
         f"{sum(len(r.get('fills') or []) for r in mkt_legs)} lots", None, None),
        ("    market sell notional",
         sum(r["notional"] for r in mkt_legs), None),
        ("", None, None),
        ("EMBEDDED GAINS SHIELDED (in-kind transfer)", shielded, LIGHT),
        (f"    across {len(ink_legs)} in-kind legs / "
         f"{sum(len(r.get('fills') or []) for r in ink_legs)} lots", None, None),
        ("    in-kind notional",
         sum(r["notional"] for r in ink_legs), None),
        ("", None, None),
        ("Net realized P&L (harvested losses only)", harvested, None),
        ("Gains shielded are NOT realized — they transfer with the shares.",
         None, None),
    ]
    rr = 3
    for label, val, color in summary_rows:
        if not label:
            rr += 1
            continue
        cl = wss.cell(row=rr, column=1, value=label)
        cl.font = Font(bold=not label.startswith(" "))
        cl.alignment = Alignment(horizontal="left")
        if val is not None:
            cv = wss.cell(row=rr, column=2, value=float(val))
            cv.number_format = FMT_USD
            cv.alignment = Alignment(horizontal="right")
            if color:
                cv.fill = PatternFill("solid", fgColor=color)
                cv.font = Font(bold=True)
        rr += 1
    wss.column_dimensions["A"].width = 52
    wss.column_dimensions["B"].width = 18

    ws = wb.create_sheet("Blotter (Expandable)")
    ws.sheet_view.showGridLines = False
    # Detail rows sit BELOW their summary (parent) row.
    ws.sheet_properties.outlinePr.summaryBelow = True

    cols = ["Security / Lot", "CUSIP", "Class", "Current", "Target",
            "Side", "Method", "Shares", "Price", "Notional", "Basis",
            "G/L per Share", "Realized / Shielded", "Open Date", "Lots", "Flags"]
    ncol = len(cols)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    ws["A1"] = (f"NXTI REBALANCE - EXPANDABLE BLOTTER  ({config.RUN_DATE})   "
                f"|  click the + buttons on the left to reveal each leg's tax lots")
    ws["A1"].font = title_font; ws["A1"].fill = title_fill; ws["A1"].alignment = center
    ws.row_dimensions[1].height = 24

    for j, h in enumerate(cols, start=1):
        c = ws.cell(row=2, column=j, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = border

    fmt = {4: FMT_SHARES, 5: FMT_SHARES, 8: FMT_SHARES, 9: FMT_PRICE,
           10: FMT_USD, 11: FMT_PRICE, 12: FMT_PRICE, 13: FMT_USD}

    def style(rng_row, fill):
        for j in range(1, ncol + 1):
            cc = ws.cell(row=rng_row, column=j)
            cc.border = border
            cc.fill = fill
            if j in fmt:
                cc.number_format = fmt[j]

    r = 3
    for rec in records:
        side, method = rec["side"], rec["method"]
        # ----- PARENT (leg) row -----
        parent = [
            rec["security_name"], rec["cusip"], rec["classification"],
            rec["current_shares"], rec["target_shares"], side, method,
            rec["shares"], rec["current_price"], rec["notional"],
            rec["avg_basis"] if not _isnan(rec["avg_basis"]) else None,
            None, rec["realized_gain_loss"], None, rec["lots_used"],
            rec["flags"] or None,
        ]
        for j, v in enumerate(parent, start=1):
            ws.cell(row=r, column=j, value=v)
        if side == "BUY":
            pfill = PatternFill("solid", fgColor=LIGHT)
        elif method == "MARKET":
            pfill = PatternFill("solid", fgColor=GREEN)
        elif method == "IN_KIND":
            pfill = PatternFill("solid", fgColor=GREY)
        elif method == "NO_TRADE":
            pfill = PatternFill("solid", fgColor="FFFFFF")
        else:
            pfill = PatternFill("solid", fgColor=AMBER)
        style(r, pfill)
        ws.cell(row=r, column=1).font = (
            italic if method == "NO_TRADE" else Font(bold=True))
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")
        parent_row = r
        r += 1

        # ----- CHILD (lot) rows, grouped + collapsed -----
        # Lots are listed in the exact order the engine consumed them. Under
        # all_or_nothing that is ALWAYS lowest cost basis first; under the
        # legacy split policy market legs run highest-cost-first instead.
        fills = rec.get("fills") or []
        n = len(fills)
        if config.SELL_METHOD_POLICY == "all_or_nothing":
            order_note = "lowest cost first" if method in ("MARKET", "IN_KIND") else ""
        else:
            order_note = ("highest cost first" if method == "MARKET"
                          else "lowest cost first" if method == "IN_KIND" else "")
        for idx, f in enumerate(fills, start=1):
            gl_share = rec["current_price"] - f.basis        # +gain / -loss per share
            lot_pnl = gl_share * f.shares
            primary = (idx == 1)
            label = (f"        ↳ lot {idx} of {n}  ·  opened {f.open_date}"
                     + ("  (SPLIT — part of lot)" if f.was_split else ""))
            note_bits = []
            if primary and n > 1:
                note_bits.append(f"PRIMARY ({order_note})")
            note_bits.append("loss realized" if method == "MARKET"
                             else "gain shielded" if method == "IN_KIND" else "")
            if f.was_split:
                note_bits.append("split lot")
            child = [
                label, None, None, None, None, None, method, f.shares,
                rec["current_price"], f.shares * rec["current_price"],
                f.basis, gl_share, lot_pnl, f.open_date, None,
                "; ".join(b for b in note_bits if b),
            ]
            for j, v in enumerate(child, start=1):
                ws.cell(row=r, column=j, value=v)
            # Highlight the primary (highest-loss / largest-gain) lot in each leg.
            cfill = PatternFill("solid", fgColor="FFF3CC" if primary and n > 1
                                else "F2F7FF")
            style(r, cfill)
            ws.cell(row=r, column=1).font = (
                Font(bold=True, color="7F6000") if primary and n > 1 else italic
            )
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", indent=2)
            ws.row_dimensions[r].outline_level = 1
            ws.row_dimensions[r].hidden = True       # start collapsed
            r += 1
        if fills:
            ws.row_dimensions[parent_row].collapsed = True

    ws.freeze_panes = "A3"
    widths = [30, 12, 11, 9, 9, 6, 10, 11, 11, 14, 11, 13, 17, 13, 6, 26]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # Legend
    lr = r + 1
    ws.cell(row=lr, column=1, value="Legend:").font = Font(bold=True)
    legend = [
        ("Buy", LIGHT), ("Market sell (loss harvested)", GREEN),
        ("In-kind (gain shielded)", GREY), ("No trade (at target)", "FFFFFF"),
        ("Needs attention", AMBER),
    ]
    for i, (lab, color) in enumerate(legend):
        cc = ws.cell(row=lr + 1 + i, column=1, value=lab)
        cc.fill = PatternFill("solid", fgColor=color)
        cc.border = border
    cc = ws.cell(row=lr + 1 + len(legend), column=1,
                 value=("Highlighted lot = PRIMARY (first consumed: lowest cost "
                        "basis first)"
                        if config.SELL_METHOD_POLICY == "all_or_nothing" else
                        "Highlighted lot = PRIMARY (first consumed: highest cost "
                        "for market sells, lowest cost for in-kind)"))
    cc.fill = PatternFill("solid", fgColor="FFF3CC"); cc.border = border

    return _save(wb, path)


def _isnan(x) -> bool:
    try:
        return bool(pd.isna(x))
    except Exception:
        return False
